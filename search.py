from dotenv import load_dotenv
import os
import openai
from openpyxl import Workbook, load_workbook



# carrega as variáveis do arquivo .env
load_dotenv()

class Search:

    def __init__(self, product):
        self.product = product




    def get_description(self, planilha):
        openai.api_key = os.getenv("OPENAI_API_KEY")
        response = openai.Completion.create(
            model="text-davinci-003",
            prompt=f"Faça uma descrição completa deste produto e coloque sua ficha técnica: {self.product}",
            temperature=0,
            max_tokens=500,
            frequency_penalty=0.0,
            presence_penalty=0.0,
        )
        filename = f'./src/{planilha}'
        print(f"filename -> {filename}")
        wb = load_workbook(filename=filename)
        ws = wb.active
        produto = response.choices[0].text
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=1).value = self.product
        ws.cell(row=proxima_linha, column=2).value = response.choices[0].text
        wb.save(filename=filename)
        return produto


    def __str__(self):
        return self.product