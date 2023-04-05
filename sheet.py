import calendar
import time
from openpyxl import Workbook, load_workbook

class sheet:

    def __init__(self):
        current_GMT = time.gmtime()
        time_stamp = calendar.timegm(current_GMT)
        self.path = 'src/'
        self.name = f"{time_stamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Nome do Produto'
        ws['B1'] = 'Descrição'
        wb.save(filename=f"{self.path}{self.name}")


    def __str__(self):
        return self.name
